from django.conf import settings
from django.contrib.gis.geos import GEOSGeometry
from django.core import mail
from django.core.mail import EmailMultiAlternatives
from django.http import FileResponse, JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render #, redirect
from django.views.generic import View

import codecs, csv, datetime, sys, openpyxl, os, pprint, re, time, logging
import pandas as pd
import simplejson as json
from chardet import detect
from django_celery_results.models import TaskResult
from shapely import wkt

from areas.models import Country
from datasets.models import Dataset, DatasetUser, Hit
from datasets.static.hashes import aat, parents, aat_q
from datasets.static.hashes import aliases as al
from .exceptions import LPFValidationError, DelimValidationError
from main.models import Log
from places.models import PlaceGeom, Type
pp = pprint.PrettyPrinter(indent=1)

"""
	TODO: broken? not in use?
  email various, incl. Celery down notice
  to ['whgazetteer@gmail.com','karl@kgeographer.org'],
"""
def emailer(subj, msg, from_addr, to_addr):
  print('subj, msg, from_addr, to_addr',subj, msg, from_addr, to_addr)
  send_mail(
      subj, msg, from_addr, to_addr,
      fail_silently=False,
  )


# ***
# DOWNLOAD FILES
# ***
# TODO: use DRF serializer? download_{format} methods on api.PlaceList() view?
# https://stackoverflow.com/questions/38697529/how-to-return-generated-file-download-with-django-rest-framework

# one-off download lp7 example tsv
# forces text/plain content_type!?
def downloadLP7(request):
	file = open('static/files/lp7_100.tsv', 'r')  # Open the specified file
	response = HttpResponse(file)  # Give file handle to HttpResponse object
	# Set the header to tell the browser that this is a file
	response['Content-Type'] = 'text/plain'
	# This is a simple description of the file. Note that the writing is the fixed one
	response['Content-Disposition'] = 'attachment;filename="lp7_100.tsv"'
	return response

# initiate celery tasks for dataset downloads
# TODO: add download Collection capability Apr 2022
def downloader(request, *args, **kwargs):
	user = request.user
	print('request.user', request.user)
	print('downloader() request.POST', request.POST)
	dsid = request.POST.get('dsid') or None
	collid = request.POST.get('collid') or None
	format = request.POST.get('format') or None
	notes = True if request.POST.get('notes') == 'true' else False

	from datasets.tasks import make_download
	# POST *should* be the only case...
	if request.method == 'POST' and request.is_ajax:
		print('ajax == True')
		# download_task = make_download(
		download_task = make_download.delay(
			{"username":user.username, "userid":user.id},
			dsid=dsid,
			collid=collid,
			notes=notes,
			format=format,
		)
		obj={'task_id':download_task.task_id}

		return HttpResponse(json.dumps(obj), content_type='application/json')

	elif request.method == 'POST' and not request.is_ajax:
		print('request.POST (not ajax)', request.POST)

	elif request.method == 'GET':
		print('request.GET', request.GET)


""" deprecatING (still used from collection download modal ) """
def download_augmented(request, *args, **kwargs):
	from django.db import connection
	print('download_augmented kwargs',kwargs)
	print('download_augmented request',request)
	username = request.user.username
	ds=get_object_or_404(Dataset,pk=kwargs['id'])
	dslabel = ds.label
	url_prefix='http://whgazetteer.org/api/place/'
	fileobj = ds.files.all().order_by('-rev')[0]
	date=makeNow()

	req_format = kwargs['format']
	if req_format is not None:
		print('download format',req_format)

	features=ds.places.all().order_by('id')

	print('download_augmented() file format', fileobj.format)
	print('download_augmented() req. format', req_format)
	start = datetime.datetime.now()
	if fileobj.format == 'delimited' and req_format in ['tsv', 'delimited']:
		# get header
		header = ds.files.all().order_by('id')[0].header
		print('making a tsv file')
		# make file name
		#fn = 'media/user_'+user+'/'+ds.label+'_aug_'+date+'.tsv'
		fn = 'media/downloads/'+username+'_'+dslabel+'_'+date+'.tsv'

		def augLinks(linklist):
			aug_links = []
			for l in linklist:
				aug_links.append(l.jsonb['identifier'])
			return ';'.join(aug_links)

		def augGeom(qs_geoms):
			gobj = {'new':[]}
			for g in qs_geoms:
				if not g.task_id:
					# it's an original
					gobj['lonlat'] = g.jsonb['coordinates']
				else:
					# it's an aug/add
					gobj['new'].append({"id":g.jsonb['citation']['id'],"coordinates":g.jsonb['coordinates'][0]})
			return gobj

		# TODO: return valid LP-TSV, incl. geowkt where applic.
		with open(fn, 'w', newline='', encoding='utf-8') as csvfile:
			writer = csv.writer(csvfile, delimiter='\t', quotechar='', quoting=csv.QUOTE_NONE)
			writer.writerow(['id','whg_pid','title','ccodes','lon','lat','added','matches'])
			#writer.writerow(header)
			for f in features:
				geoms = f.geoms.all()
				gobj = augGeom(geoms)
				#print('gobj',f.id, gobj)
				row = [str(f.src_id),
				       str(f.id),
				       f.title,
				       ';'.join(f.ccodes),
				       gobj['lonlat'][0] if 'lonlat' in gobj else None,
				       gobj['lonlat'][1] if 'lonlat' in gobj else None,
				       gobj['new'] if 'new' in gobj else None,
				       str(augLinks(f.links.all()))
				       ]
				writer.writerow(row)
				#progress_recorder.set_progress(i + 1, len(features), description="tsv progress")
		response = FileResponse(open(fn, 'rb'), content_type='text/csv')
		response['Content-Disposition'] = 'attachment; filename="'+os.path.basename(fn)+'"'
		end = datetime.datetime.now()
		print('elapsed tsv', end-start)
		return response
	else:
		print('building lpf file')
		# make file name
		fn = 'media/downloads/'+username+'_'+dslabel+'_'+date+'.tsv'
		result={"type":"FeatureCollection","features":[],
		        "@context": "https://raw.githubusercontent.com/LinkedPasts/linked-places/master/linkedplaces-context-v1.1.jsonld",
		        "filename": "/"+fn}
		print('augmented lpf template', result)
		with open(fn, 'w', encoding='utf-8') as outfile:
			with connection.cursor() as cursor:
				cursor.execute("""with namings as
          (select place_id, jsonb_agg(jsonb) as names from place_name pn
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          placetypes as
          (select place_id, jsonb_agg(jsonb) as "types" from place_type pt
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          placelinks as
          (select place_id, jsonb_agg(jsonb) as links from place_link pl
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          geometry as
          (select place_id, jsonb_agg(jsonb) as geoms from place_geom pg
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          placewhens as
          (select place_id, jsonb as whenobj from place_when pw
          where place_id in (select id from places where dataset = '{ds}')),
          placerelated as
          (select place_id, jsonb_agg(jsonb) as rels from place_related pr
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          descriptions as
          (select place_id, jsonb_agg(jsonb) as descrips from place_description pdes
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id ),
          depictions as
          (select place_id, jsonb_agg(jsonb) as depicts from place_depiction pdep
          where place_id in (select id from places where dataset = '{ds}')
          group by place_id )
          select jsonb_build_object(
            'type','Feature',
            '@id', p.src_id,
            'properties', jsonb_build_object(
                'pid', '{urlpre}'||p.id,
                'title', p.title),
            'names', n.names,
            'types', coalesce(pt.types, '[]'),
            'links', coalesce(pl.links, '[]'),
            'geometry', case when g.geoms is not null
                then jsonb_build_object(
                'type','GeometryCollection',
                'geometries', g.geoms)
                else jsonb_build_object(
                'type','Point','coordinates','{a}'::char[])
                end,
            'when', pw.whenobj,
            'relations',coalesce(pr.rels, '[]'),
            'descriptions',coalesce(pdes.descrips, '[]'),
            'depictions',coalesce(pdep.depicts, '[]')
          ) from places p
          left join namings n on p.id = n.place_id
          left join placetypes pt on p.id = pt.place_id
          left join placelinks pl on p.id = pl.place_id
          left join geometry g on p.id = g.place_id
          left join placewhens pw on p.id = pw.place_id
          left join placerelated pr on p.id = pr.place_id
          left join descriptions pdes on p.id = pdes.place_id
          left join depictions pdep on p.id = pdep.place_id
          where dataset = '{ds}'
        """.format(urlpre=url_prefix, ds=dslabel, a='{}'))
				for row in cursor:
					g = row[0]['geometry']
					# get rid of empty/unknown geometry
					if g['type'] != 'GeometryCollection' and g['coordinates'] == []:
						row[0].pop('geometry')
					result['features'].append(row[0])
					#progress_recorder.set_progress(i + 1, len(features), description="lpf progress")
				outfile.write(json.dumps(result,indent=2))
				#outfile.write(json.dumps(result))

		end = datetime.datetime.now()
		print('elapsed lpf', end-start)
		# response is reopened file
		response = FileResponse(open(fn, 'rb'), content_type='text/json')
		response['Content-Disposition'] = 'attachment; filename="'+os.path.basename(fn)+'"'

		return response

""" just gets a file and downloads it to File/Save window """
def download_file(request, *args, **kwargs):
	ds=get_object_or_404(Dataset,pk=kwargs['id'])
	fileobj = ds.files.all().order_by('-rev')[0]
	fn = 'media/'+fileobj.file.name
	file_handle = fileobj.file.open()
	print('download_file: kwargs,fn,fileobj.format',kwargs,fn,fileobj.format)
	# set content type
	response = FileResponse(file_handle, content_type='text/csv' if fileobj.format=='delimited' else 'text/json')
	response['Content-Disposition'] = 'attachment; filename="'+fileobj.file.name+'"'

	return response

#
# experiment (deprecated?)
def download_augmented_slow(request, *args, **kwargs):
	print('download_augmented kwargs',kwargs)
	user = request.user.username
	ds=get_object_or_404(Dataset,pk=kwargs['id'])
	fileobj = ds.files.all().order_by('-rev')[0]
	date=makeNow()

	req_format = kwargs['format']
	if req_format is not None:
		print('got format',req_format)
		#qs = qs.filter(title__icontains=query)

	features=ds.places.all()

	if fileobj.format == 'delimited' and req_format == 'tsv':
		print('augmented for delimited')
		# make file name
		fn = 'media/user_'+user+'/'+ds.label+'_aug_'+date+'.tsv'
		def augLinks(linklist):
			aug_links = []
			for l in linklist:
				aug_links.append(l.jsonb['identifier'])
			return ';'.join(aug_links)

		def augGeom(qs_geoms):
			gobj = {'new':[]}
			for g in qs_geoms:
				if not g.task_id:
					# it's an original
					gobj['lonlat'] = g.jsonb['coordinates']
				else:
					# it's an aug/add
					gobj['new'].append({"id":g.jsonb['citation']['id'],"coordinates":g.jsonb['coordinates']})
			return gobj

		with open(fn, 'w', newline='', encoding='utf-8') as csvfile:
			writer = csv.writer(csvfile, delimiter='\t', quotechar='', quoting=csv.QUOTE_NONE)
			writer.writerow(['id','whg_pid','title','ccodes','lon','lat','added','matches'])
			for f in features:
				geoms = f.geoms.all()
				gobj = augGeom(geoms)
				row = [
					str(f.src_id),
					str(f.id),f.title,
					';'.join(f.ccodes),
					gobj['lonlat'][0] if 'lonlat' in gobj else None,
					gobj['lonlat'][1] if 'lonlat' in gobj else None,
					gobj['new'] if 'new' in gobj else None,
					str(augLinks(f.links.all())) ]
				writer.writerow(row)
				#print(row)
		response = FileResponse(open(fn, 'rb'),content_type='text/csv')
		response['Content-Disposition'] = 'attachment; filename="'+os.path.basename(fn)+'"'

		return response
	else:
		# make file name
		fn = 'media/user_'+user+'/'+ds.label+'_aug_'+date+'.json'

		with open(fn, 'w', encoding='utf-8') as outfile:
			#fcoll = {"type":"FeatureCollection","features":[]}
			for f in features:
				print('dl_aug, lpf adding feature:',f)
				feat={"type":"Feature",
				      "properties":{
					      "@id":f.dataset.uri_base+f.src_id,
					      "src_id":f.src_id,
					      "title":f.title,
					      "whg_pid":f.id}}
				if len(f.geoms.all()) >1:
					feat['geometry'] = {'type':'GeometryCollection'}
					feat['geometry']['geometries'] = [g.jsonb for g in f.geoms.all()]
				elif len(f.geoms.all()) == 1:
					feat['geometry'] = f.geoms.first().jsonb
				else: # no geoms
					feat['geometry'] = feat['geometry'] = {'type':'GeometryCollection','geometries':[]}
				feat['names'] = [n.jsonb for n in f.names.all()]
				feat['types'] = [t.jsonb for t in f.types.all()]
				feat['when'] = [w.jsonb for w in f.whens.all()]
				feat['relations'] = [r.jsonb for r in f.related.all()]
				feat['links'] = [l.jsonb for l in f.links.all()]
				feat['descriptions'] = [des.jsonb for des in f.descriptions.all()]
				feat['depictions'] = [dep.jsonb for dep in f.depictions.all()]
				#fcoll['features'].append(feat)
				outfile.write(json.dumps(feat,indent=2))
			#outfile.write(json.dumps(fcoll,indent=2))

		# response is reopened file
		response = FileResponse(open(fn, 'rb'), content_type='text/json')
		#response['Content-Disposition'] = 'attachment; filename="'+os.path.basename(fn)+'"'
		response['Content-Disposition'] = 'filename="'+os.path.basename(fn)+'"'

		return response
	# *** /end DOWNLOAD FILES

# GeoJSON for all places in a dataset
# feeds ds_browse (owner view); ds_places, collection_places (public)
def fetch_geojson_ds(request, *args, **kwargs):
	# print('fetch_geojson_ds kwargs',kwargs)
	dsid=kwargs['dsid']
	ds=get_object_or_404(Dataset,pk=dsid)

	# build a fast FeatureCollection
	features=PlaceGeom.objects.filter(place_id__in=ds.placeids).values_list(
		'jsonb','place_id','src_id','place__title','place__minmax', 'place__fclasses')
	fcoll = {"type":"FeatureCollection","features":[], "minmax":ds.minmax}
	for f in features:
		# some places have no temporal scoping (dplace, geonames, etc.)
		minmax = f[4] if f[4] and len(f[4]) == 2 else None
		feat={"type":"Feature",
		      "properties":{"pid":f[1],"src_id":f[2],"title":f[3],
		                    "fclasses":f[5], "ds":ds.label},
		      "geometry":f[0]}
		if minmax:
			feat["properties"]["min"] = minmax[0]
			feat["properties"]["max"] = minmax[1]
		fcoll['features'].append(feat)

	result = {"minmax":ds.minmax, "collection":fcoll}
	return JsonResponse(result, safe=False,json_dumps_params={'ensure_ascii':False})

# flatten for gl time-mapping
# one feature per geometry w/min & max
def fetch_geojson_flat(request, *args, **kwargs):
	print('fetch_geojson_flat kwargs',kwargs)
	dsid=kwargs['dsid']
	ds=get_object_or_404(Dataset,pk=dsid)

	#from places.models import PlaceGeom
	#from datasets.models import Dataset
	#ds= Dataset.objects.get(pk=1106)

	fcoll = {"type":"FeatureCollection","features":[]}

	# build a FLAT FeatureCollection
	pgobjects=PlaceGeom.objects.filter(place_id__in=ds.placeids)
	#pgobjects=PlaceGeom.objects.filter(place_id__in=ds.placeids).values_list(
	#'jsonb','place_id','src_id','minmax','place__title','place__minmax', 'place__fclasses')
	for pg in pgobjects:
		geom = json.loads(pg.geom.geojson)
		if geom['type'] != 'GeometryCollection':
			fcoll['features'].append({"type":"Feature",
			                          "geometry":geom,
			                          "properties":{
				                          "id":pg.place_id, "title":pg.place.title,
				                          "min":pg.minmax[0] if pg.minmax else None,
				                          "max":pg.minmax[1] if pg.minmax else None }}
			                         )

	result = {"minmax":ds.minmax, "collection":fcoll}
	return JsonResponse(result, safe=False,json_dumps_params={'ensure_ascii':False})

def get_encoding_excel(fn):
	fin = codecs.open(fn, 'r')
	encoding = fin.encoding
	fin.close()
	return encoding

def get_encoding_delim(fn):
	with open(fn, 'rb') as f:
		rawdata = f.read()
	# print('detect', detect(rawdata))
	return detect(rawdata)['encoding']

def groom_files(user):
	return 'groomed files for ' + user

	# ***
# format validation errors for display
# ***
def parse_errors_tsv(errors):
	new_errors = []
	for e in errors:
		newe = re.sub('a constraint: constraint', 'the constraint:', e)
		newe = re.sub('at position "(\\d+)" does', 'does', newe)
		newe = re.sub('row at position "(\\d+)"', 'row \\1', newe)
		newe = re.sub('value', 'cell', newe)
		new_errors.append(newe)
	return new_errors
# tsv error: ['The cell "" in row at position "2" and field "id" at position "1" does not conform to a constraint: constraint "required" is "True"', 'The row at position "2" does not conform to the primary key constraint: cells composing the primary keys are all "None"']

# lpf error: [{'feat': 2, 'error': "'null' is not of type 'object'"}, {'feat': 3, 'error': "'null' is not of type 'object'"}]

# *** NOT YET CALLED
# format lpf validation errors for modal display
# ***
#def parse_errors_lpf(errors):
#new_errors = []
#for e in error
#print('parse_errors_lpf()',errors)

# ***
# parse start & end for insert to db
# TODO: is year-only minmax useful in GUI?
# parsedates for ds_insert_lpf will be different
# ***
#def intmap(arr):
#return [int(a) for a in arr]

def parse_errors_lpf(errors):
	print('relative_path 0',errors[0]['error'].relative_path)
	"deque(['geometry', 'geometries', 0])"
	msg = [{"row":e['feat'], "msg":e['error'].message, "path":
		re.search('deque\(\[(.*)\]\)',
		          str(e['error'].relative_path)).group(1) } for e in errors]
	return msg

#
# called by ds_insert_tsv()
# returns object for PlaceWhen.jsonb in db
# and minmax int years for PlacePortalView()
# TODO: patched Apr 2023; needs refactor
def parsedates_tsv(dates):
	# print('parsedates() dates', dates)
	# they are either a string or None at this point
	s_yr=int(dates[0]) if dates[0] else None  #; print('s_yr:', s_yr)
	e_yr=int(dates[1]) if dates[1] else None  #; print('e_yr:', e_yr)
	att_yr=int(dates[2]) if dates[2] else None  #; print('src_yr:', att_yr)
	result = {
		"timespans":[{
			"start": {
				"earliest":s_yr or None},
			"end": {
				"latest":e_yr or None}
		}],
		"minmax":[s_yr or -99999, e_yr or 9999],
		"attestation_year": att_yr
	}
	# print('parsedates_tsv() result', result)
	return result

# extract integers for new Place from lpf
def timespansReduce(tsl):
	result = []
	for ts in tsl:
		s = ts['start'][list(ts['start'].keys())[0]]
		s_yr=s[:5] if s[0] == '-' else s[:4]
		#e = ts['end'][list(ts['end'].keys())[0]] \
		#if 'end' in ts else s
		# lpf imports from tsv exports can have '' in end
		end = ts['end'][list(ts['end'].keys())[0]] if 'end' in ts else None # no end
		e = end if end and end != '' else s
		e_yr=e[:5] if e[0] == '-' else e[:4]
		result.append([int(s_yr), int(e_yr)])
		#s = int(ts['start'][list(ts['start'].keys())[0]])
		#e = int(ts['end'][list(ts['end'].keys())[0]]) \
		#if 'end' in ts else s
		#result.append([s,e])
	return result

#
# called by ds_insert_lpf()
# TODO: replicate outcome of parsedates_tsv()
#
def parsedates_lpf(feat):
	intervals=[]
	# gather all when elements
	# global when?
	if 'when' in feat and 'timespans' in feat['when']:
		try:
			intervals += timespansReduce(feat['when']['timespans'])
		except:
			print('parsedates_lpf hung on', feat['@id'])

	# which feat keys might have a when?
	possible_keys = list(set(feat.keys() & \
	                         set(['names','types','relations','geometry'])))
	print('possible_keys in parsedates_lpf()',possible_keys)

	# first, geometry
	# collections...
	geom = feat['geometry'] if 'geometry' in feat else None
	if geom and geom['type'] == 'GeometryCollection':
		for g in geom['geometries']:
			if 'when' in g and 'timespans' in g['when']:
				intervals += timespansReduce(g['when']['timespans'])
	# or singleton
	else:
		if geom and 'when' in geom:
			if 'timespans' in geom['when']:
				intervals += timespansReduce(g['when']['timespans'])

	# then the rest
	for k in possible_keys:
		if k != 'geometry':
			obj = feat[k]
			for o in obj:
				if 'when' in o and 'timespans' in o['when']:
					intervals += timespansReduce(o['when']['timespans'])
					# features must have >=1 when, with >=1 timespan
	# absent end replaced by start by timespansReduce()
	starts = [ts[0] for ts in intervals]
	ends = [ts[1] for ts in intervals]
	# some lpf records have no time at all b/c not required as with LP-TSV
	minmax = [
		int(min(starts)) if len(starts)>0 else None,
		int(max(ends))  if len(ends)>0 else None
	]
	# de-duplicate
	unique=list(set(tuple(sorted(sub)) for sub in intervals))
	return {"intervals": unique, "minmax": minmax}

def parse_validation_error(error):
    # Extract key parts of the error message
    # print('parse_validation_error()', error)
    data = error.instance
    message = error.message
    schema_path = " -> ".join([str(p) for p in error.absolute_schema_path])

    # Construct a user-friendly message
    user_message = f"Error in data: {data}. Reason: {message}. Schema path: {schema_path}"

    return user_message

#
# validate Linked Places json-ld (w/jsonschema)
# format ['coll' (FeatureCollection) | 'lines' (json-lines)]
# TODO: 'format' will eventually support jsonlines
def validate_lpf(tempfn, format):
	logger = logging.getLogger('django')
	# print('in validate_lpf()...format', format)
	schema = json.loads(codecs.open('datasets/static/validate/schema_lpf_v1.2.json','r','utf8').read())

	# rename tempfn
	newfn = tempfn+'.jsonld'
	os.rename(tempfn,newfn)
	infile = codecs.open(newfn, 'r', 'utf8')
	result = {"format":"lpf", "errors":[]}

	# TODO: handle json-lines
	jdata = json.loads(infile.read())
	if len(set(['type', '@context', 'features'])-set(jdata.keys())) > 0 \
		or jdata['type'] != 'FeatureCollection' \
		or len(jdata['features']) == 0:
		print('not valid GeoJSON-LD')
	else:
		errors = []
		seen_error_paths = set()

		for countrows, feat in enumerate(jdata['features'], start=1):

			if len(errors) >= 3:  # Stop after collecting 3 errors
				break
			try:
				validate(
					instance=feat,
					schema=schema,
					format_checker=draft7_format_checker
				)
			except ValidationError as e:
				error_path = " -> ".join([str(p) for p in e.absolute_path])
				if error_path not in seen_error_paths:  # Check if this error type (path) has been seen before
					detailed_error = parse_validation_error(e)
					errors.append({"feat": countrows, 'error': detailed_error})
					seen_error_paths.add(error_path)

		print('errors in validate_lpf()', errors)
		if errors:
			aggregated_message = "; ".join([error['error'] for error in errors])
			if len(errors) == 3:
				aggregated_message += " ... Your uploaded file has more errors; these were the first three found."
			# print('aggregated_message',aggregated_message )
			raise LPFValidationError(errors)
			# raise LPFValidationError(aggregated_message)

		result['count'] = countrows
	return result

def validate_delim(df):
	errors = []

	# Define required fields and patterns
	required_fields = ['id', 'title', 'title_source', 'start']
	pattern_constraints = {
		'ccodes': "([a-zA-Z]{2};?)+",
		'matches': "(https?:\\/\\/.*\\..*;?)+|([a-z]{1,8}:.*;?)+",
		'parent_id': "(https?:\/\/.*\\..*|#\\d*)",
		'start': "(-?\\d{1,4}(-\\d{2})?(-\\d{2})?)(\/(-?\\d{1,4}(-\\d{2})?(-\\d{2})?))?",
		'end': "(-?\\d{1,4}(-\\d{2})?(-\\d{2})?)(\/(-?\\d{1,4}(-\\d{2})?(-\\d{2})?))?"
	}
	range_constraints = {
		'lon': (-180, 180),
		'lat': (-90, 90)
	}

	# Loop through rows for validation
	for index, row in df.iterrows():
		# Check required fields
		for field in required_fields:
			if field not in row:
				errors.append({"row": index + 1, "error": f"Required field missing: {field}"})

		# Check for either "parent_name" or "parent_id"
		if not ("parent_name" in row or "parent_id" in row):
			errors.append({"row": index + 1, "error": "Either 'parent_name' or 'parent_id' must be present"})

		# Check pattern constraints
		for field, pattern in pattern_constraints.items():
			if field in row and not bool(re.search(pattern, str(row[field]))):
				errors.append(
					{"row": index + 1, "error": f"Field {field} contains a value that does not match the required pattern"})

		# Check range constraints
		for field, (low, high) in range_constraints.items():
			if field in row and (row[field] < low or row[field] > high):
				errors.append({"row": index + 1, "error": f"Value in {field} is out of the allowed range"})

	if errors:
		raise DelimValidationError(errors)

	return errors


#
# validate LP-TSV file (uses frictionless.py 3.31.0)
#
def validate_tsv(fn, ext):
	# incoming csv or tsv; in cases converted from xlsx or ods via pandas
	# print('validate_tsv() fn', fn)
	# pull header for missing columns test below
	header = codecs.open(fn, 'r').readlines()[0][:-1]
	header = list(map(str.lower, header.split('\t' if '\t' in header else ',')))
	# header = header.split('\t' if '\t' in header else ',')
	list(map(str.lower, header))
	# print('header', header)
	result = {"format":"delimited", "errors":[], "columns":header}
	schema_lptsv = json.loads(codecs.open('datasets/static/validate/schema_tsv.json', 'r', 'utf8').read())
	try:
		report = fvalidate(fn, schema=schema_lptsv, sync_schema=True)
	except:
		err = sys.exc_info()
		result['errors'].append('File failed format validation. Error: '+err+'; '+str(err[1].args))
		print('error on fvalidate',err)
		print('error args',err[1].args)
		return result

	if len(report['tables']) > 0:
		rpt = report['tables'][0]
		result['count'] = rpt['stats']['rows']  # count
		print('rpt errors', rpt['errors'])

	req = ['id', 'title', 'title_source']
	missing = list(set(req) - set(header))

	# filter harmless errors
	# TODO: is filtering encoding-error here problematic?
	result['errors'] = [x['message'] for x in rpt['errors'] \
	                    if x['code'] not in ["blank-header", "missing-header", "encoding-error"]]
	if len(missing) > 0:
		result['errors'].insert(0,'Required column(s) missing or header malformed: '+
		                        ', '.join(missing))

	return result

class HitRecord(object):
	def __init__(self, place_id, dataset, src_id, title):
		self.place_id = place_id
		self.src_id = src_id
		self.title = title
		self.dataset = dataset

	def __str__(self):
		import json
		return json.dumps(str(self.__dict__))

	def toJSON(self):
		import json
		return json.loads(json.dumps(self.__dict__,indent=2))


class PlaceMapper(object):
	def __init__(self, id, src_id, title):
		self.id = id
		self.src_id = src_id
		self.title = title

	def __setitem__(self, key, value):
		setattr(self, key, value)

	def __getitem__(self, key):
		return getattr(self, key)

	def __str__(self):
		import json
		return json.dumps(str(self.__dict__))

	def toJSON(self):
		import json
		return json.loads(json.dumps(self.__dict__, indent=2))

def is_aat(string):
	return True if string.startswith('aat') or 'aat/' in string else False

# null fclass: 300239103, 300056006, 300155846
# refactored to use Type model in db
def aat_lookup(aid):
	try:
		typeobj = get_object_or_404(Type, aat_id=aid)
		return typeobj.term
	except:
		print(str(aid)+' broke aat_lookup()', sys.exc_info())
		# return {"label": None, "fclass":None}
		return None

#u='https://catalogue.bnf.fr/ark:/12148/cb193409'
def aliasIt(url):
	r1=re.compile(r"\/(?:.(?!\/))+$")
	id=re.search(r1,url)
	if id:
		id = id.group(0)[1:].replace('cb','')
	r2 = re.compile(r"bnf|cerl|dbpedia|geonames|d-nb|loc|pleiades|tgn|viaf|wikidata|whg|wikipedia")
	tag=re.search(r2,url)
	if tag and id:
		return al.tags[tag.group(0)]['alias']+':'+id
	else:
		return url

# flattens nested tuple list
def flatten(l):
	for el in l:
		if isinstance(el, tuple) and any(isinstance(sub, tuple) for sub in el):
			for sub in flatten(el):
				yield sub
		else:
			yield el

def format_size(num):
	return round(num/1000000, 2)

"""
  'monkey patch' for hully() for acknowledged GEOS/Django issue
  "call this any time before using GEOS features" @bpartridge says
"""
def patch_geos_signatures():
	"""
	Patch GEOS to function on macOS arm64 and presumably
	other odd architectures by ensuring that call signatures
	are explicit, and that Django 4 bugfixes are backported.

	Should work on Django 2.2+, minimally tested, caveat emptor.
	"""
	import logging

	from ctypes import POINTER, c_uint, c_int
	from django.contrib.gis.geos import GeometryCollection, Polygon
	from django.contrib.gis.geos import prototypes as capi
	from django.contrib.gis.geos.prototypes import GEOM_PTR
	from django.contrib.gis.geos.prototypes.geom import GeomOutput
	from django.contrib.gis.geos.libgeos import geos_version, lgeos
	from django.contrib.gis.geos.linestring import LineString

	logger = logging.getLogger("geos_patch")

	_geos_version = geos_version()
	logger.debug("GEOS: %s %s", _geos_version, repr(lgeos))

	# Backport https://code.djangoproject.com/ticket/30274
	def new_linestring_iter(self):
		for i in range(len(self)):
			yield self[i]

	LineString.__iter__ = new_linestring_iter

	# macOS arm64 requires that we have explicit argtypes for cffi calls.
	# Patch in argtypes for `create_polygon` and `create_collection`,
	# and then ensure their prep functions do NOT use byref so that the
	# arrays (`(GEOM_PTR * length)(...)`) auto-convert into `Geometry**`.
	# create_empty_polygon doesn't need to be patched as it takes no args.

	# Geometry*
	# GEOSGeom_createPolygon_r(GEOSContextHandle_t extHandle,
	#   Geometry* shell, Geometry** holes, unsigned int nholes)
	capi.create_polygon = GeomOutput(
		"GEOSGeom_createPolygon", argtypes=[GEOM_PTR, POINTER(GEOM_PTR), c_uint]
	)

	# Geometry*
	# GEOSGeom_createCollection_r(GEOSContextHandle_t extHandle,
	#   int type, Geometry** geoms, unsigned int ngeoms)
	capi.create_collection = GeomOutput(
		"GEOSGeom_createCollection", argtypes=[c_int, POINTER(GEOM_PTR), c_uint]
	)

	# The below implementations are taken directly from Django 2.2.25 source;
	# the only changes are unwrapping calls to byref().

	def new_create_polygon(self, length, items):
		# Instantiate LinearRing objects if necessary, but don't clone them yet
		# _construct_ring will throw a TypeError if a parameter isn't a valid ring
		# If we cloned the pointers here, we wouldn't be able to clean up
		# in case of error.
		if not length:
			return capi.create_empty_polygon()

		rings = []
		for r in items:
			if isinstance(r, GEOM_PTR):
				rings.append(r)
			else:
				rings.append(self._construct_ring(r))

		shell = self._clone(rings.pop(0))

		n_holes = length - 1
		if n_holes:
			holes = (GEOM_PTR * n_holes)(*[self._clone(r) for r in rings])
			holes_param = holes
		else:
			holes_param = None

		return capi.create_polygon(shell, holes_param, c_uint(n_holes))

	Polygon._create_polygon = new_create_polygon

	# Need to patch to not call byref so that we can cast to a pointer
	def new_create_collection(self, length, items):
		# Creating the geometry pointer array.
		geoms = (GEOM_PTR * length)(
			*[
				# this is a little sloppy, but makes life easier
				# allow GEOSGeometry types (python wrappers) or pointer types
				capi.geom_clone(getattr(g, "ptr", g))
				for g in items
			]
		)
		return capi.create_collection(c_int(self._typeid), geoms, c_uint(length))

	GeometryCollection._create_collection = new_create_collection

"""
  added patch_geos_signatures() from https://gist.github.com/bpartridge/26a11b28415d706bfb9993fc28767d68
  per https://github.com/libgeos/geos/issues/528#issuecomment-997327327
"""
def hully(g_list):
	"""
	Patch GEOS to function on macOS arm64 and presumably
	other odd architectures by ensuring that call signatures
	are explicit, and that Django 4 bugfixes are backported.

	Should work on Django 2.2+, minimally tested, caveat emptor.
	"""
	import logging

	from ctypes import POINTER, c_uint, c_int
	from django.contrib.gis.geos import GeometryCollection, Polygon
	from django.contrib.gis.geos import prototypes as capi
	from django.contrib.gis.geos.prototypes import GEOM_PTR
	from django.contrib.gis.geos.prototypes.geom import GeomOutput
	from django.contrib.gis.geos.libgeos import geos_version, lgeos
	from django.contrib.gis.geos.linestring import LineString

	logger = logging.getLogger("geos_patch")

	_geos_version = geos_version()
	logger.debug("GEOS: %s %s", _geos_version, repr(lgeos))

	# Backport https://code.djangoproject.com/ticket/30274
	def new_linestring_iter(self):
		for i in range(len(self)):
			yield self[i]

	LineString.__iter__ = new_linestring_iter

	capi.create_polygon = GeomOutput(
		"GEOSGeom_createPolygon", argtypes=[GEOM_PTR, POINTER(GEOM_PTR), c_uint]
	)

	capi.create_collection = GeomOutput(
		"GEOSGeom_createCollection", argtypes=[c_int, POINTER(GEOM_PTR), c_uint]
	)

	def new_create_polygon(self, length, items):
		if not length:
			return capi.create_empty_polygon()

		rings = []
		for r in items:
			if isinstance(r, GEOM_PTR):
				rings.append(r)
			else:
				rings.append(self._construct_ring(r))

		shell = self._clone(rings.pop(0))

		n_holes = length - 1
		if n_holes:
			holes = (GEOM_PTR * n_holes)(*[self._clone(r) for r in rings])
			holes_param = holes
		else:
			holes_param = None

		return capi.create_polygon(shell, holes_param, c_uint(n_holes))

	Polygon._create_polygon = new_create_polygon

	# Need to patch to not call byref so that we can cast to a pointer
	def new_create_collection(self, length, items):
		# Creating the geometry pointer array.
		geoms = (GEOM_PTR * length)(
			*[
				# this is a little sloppy, but makes life easier
				# allow GEOSGeometry types (python wrappers) or pointer types
				capi.geom_clone(getattr(g, "ptr", g))
				for g in items
			]
		)
		return capi.create_collection(c_int(self._typeid), geoms, c_uint(length))

	GeometryCollection._create_collection = new_create_collection

	"""end hotfix """

	# 1 point -> Point; 2 points -> LineString; >2 -> Polygon
	try:
		mp = [GEOSGeometry(json.dumps(g)) for g in g_list]
		hull=GeometryCollection(mp).convex_hull
		# hull=GeometryCollection([GEOSGeometry(json.dumps(g)) for g in g_list]).convex_hull
	except:
		print('hully() failed on g_list', g_list)

	if hull.geom_type in ['Point', 'LineString', 'Polygon']:
		# buffer hull, but only a little if near meridian
		coll = GeometryCollection([GEOSGeometry(json.dumps(g)) for g in g_list]).simplify()
		#longs = list(c[0] for c in coll.coords)
		longs = list(c[0] for c in flatten(coll.coords))
		try:
			if len([i for i in longs if i >= 175]) == 0:
				hull = hull.buffer(.4) # ~50 km 11 Apr 2023
				# hull = hull.buffer(1) #
				# hull = hull.buffer(1.4) # ~100km radius ??
				# hull = hull.buffer(2.0) # ??
			else:
				hull = hull.buffer(0.1)
		except:
			print('hully buffer error longs:', longs )
	#print(hull.geojson)
	return json.loads(hull.geojson) if hull.geojson !=None else []

def parse_wkt(g):
	print('wkt',g)
	from shapely.geometry import mapping
	gw = wkt.loads(g)
	feature = json.loads(json.dumps(mapping(gw)))
	print('wkt, feature',g, feature)
	return feature

# from timestamp
def makeDate(ts, form):
	expr = ts.strftime("%Y-%m-%d") if form == 'iso' \
		else ts.strftime("%d-%b-%Y")
	return expr
#
def makeNow(short=False):
	ts = time.time()
	sttime = datetime.datetime.fromtimestamp(ts).strftime('%Y%m%d_%H%M%S')
	if short==True:
		sttime = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
	return sttime
#
def myprojects(me):
	return DatasetUser.objects.filter(user_id_id=me.id).values_list('dataset_id_id')
#
def parsejson(value,key):
	"""returns value for given key"""
	print('parsejson() value',value)
	obj = json.loads(value.replace("'",'"'))
	return obj[key]
#
def makeCoords(lonstr,latstr):
	#print(type(lonstr),latstr)
	lon = float(lonstr) if lonstr not in ['','nan',None] else ''
	lat = float(latstr) if latstr not in ['','nan',None] else ''
	coords = [] if (lonstr == ''  or latstr == '') else [lon,lat]
	return coords

# might be GeometryCollection or singleton
def ccodesFromGeom(geom):
	if geom['type'] == 'Point' and geom['coordinates'] == []:
		ccodes = []
		return ccodes
	else:
		g = GEOSGeometry(str(geom))
		if g.geom_type == 'GeometryCollection':
			# just hull them all
			qs = Country.objects.filter(mpoly__intersects=g.convex_hull)
		else:
			qs = Country.objects.filter(mpoly__intersects=g)
		ccodes = [c.iso for c in qs]
		return ccodes
#
def elapsed(delta):
	minutes, seconds = divmod(delta.seconds, 60)
	return '{:02}:{:02}'.format(int(minutes), int(seconds))

# called from es_lookup_tgn(); applicable for tgn only
def bestParent(qobj, flag=False):
	best = []
	# merge parent country/ies & parents
	if len(qobj['countries']) > 0 and qobj['countries'][0] != '':
		for c in qobj['countries']:
			best.append(parents.ccodes[0][c.upper()]['tgnlabel'])
	if len(qobj['parents']) > 0:
		for p in qobj['parents']:
			best.append(p)
	if len(best) == 0:
		best = ['World']
	return best

# wikidata Qs from ccodes
#TODO: consolidate hashes
def getQ(arr,what):
	#print('arr,what',arr, what)
	qids=[]
	if what == 'ccodes':
		from datasets.static.hashes.parents import ccodes
		for c in arr:
			if c.upper() in ccodes[0]:
				qids.append('wd:'+ccodes[0][c.upper()]['wdid'].upper())
	elif what == 'types':
		# if no types, add human settlement
		# if len(arr) == 0:
		#   qids.append('wd:Q486972')
		for t in arr:
			if t in aat_q.qnums:
				for q in aat_q.qnums[t]:
					qids.append('wd:'+q)
			# else:
			#   if type not found in mappings
			#   qids.append('wd:Q486972')
	return list(set(qids))

def roundy(x, direct="up", base=10):
	import math
	if direct == "down":
		return int(math.ceil(x / 10.0)) * 10 - base
	else:
		return int(math.ceil(x / 10.0)) * 10

def fixName(toponym):
	import re
	search_name = toponym
	r1 = re.compile(r"(.*?), Gulf of")
	r2 = re.compile(r"(.*?), Sea of")
	r3 = re.compile(r"(.*?), Cape")
	r4 = re.compile(r"^'")
	if bool(re.search(r1,toponym)):
		search_name = "Gulf of " + re.search(r1,toponym).group(1)
	if bool(re.search(r2,toponym)):
		search_name = "Sea of " + re.search(r2,toponym).group(1)
	if bool(re.search(r3,toponym)):
		search_name = "Cape " + re.search(r3,toponym).group(1)
	if bool(re.search(r4,toponym)):
		search_name = toponym[1:]
	return search_name if search_name != toponym else toponym

# in: list of Black atlas place types
# returns list of equivalent classes or types for {gaz}
def classy(gaz, typeArray):
	import codecs, json
	#print(typeArray)
	types = []
	finhash = codecs.open('../data/feature-classes.json', 'r', 'utf8')
	classes = json.loads(finhash.read())
	finhash.close()
	if gaz == 'gn':
		t = classes['geonames']
		default = 'P'
		for k,v in t.items():
			if not set(typeArray).isdisjoint(t[k]):
				types.append(k)
			else:
				types.append(default)
	elif gaz == 'tgn':
		t = classes['tgn']
		default = 'inhabited places' # inhabited places
		# if 'settlement' exclude others
		typeArray = ['settlement'] if 'settlement' in typeArray else typeArray
		# if 'admin1' (US states) exclude others
		typeArray = ['admin1'] if 'admin1' in typeArray else typeArray
		for k,v in t.items():
			if not set(typeArray).isdisjoint(t[k]):
				types.append(k)
			else:
				types.append(default)
	elif gaz == "dbp":
		t = classes['dbpedia']
		default = 'Place'
		for k,v in t.items():
			# is any Black type in dbp array?
			# TODO: this is crap logic, fix it
			if not set(typeArray).isdisjoint(t[k]):
				types.append(k)
	if len(types) == 0:
		types.append(default)
	return list(set(types))


# log recon action & update status
def post_recon_update(ds, user, task, test):
	print('test in utils.post_recon_update()', test )
	if test == "off":
		if task == 'idx':
			ds.ds_status = 'indexed' if ds.unindexed == 0 else 'accessioning'
		else:
			ds.ds_status = 'reconciling'
		ds.save()
	else:
		task += '_test'
	# recon task has completed, log it
	logobj = Log.objects.create(
		category = 'dataset',
		logtype = 'ds_recon',
		subtype = 'align_'+task,
		dataset_id = ds.id,
		user_id = user.id
	)
	logobj.save()
	# print('post_recon_update() logobj',logobj)

def status_emailer(ds, task_name):
	try:
		tasklabel = 'Wikidata' if task_name=='wd' else 'WHG index'
		text_content="Greetings! A "+tasklabel+" reconciliation task for the dataset "+ds.title+" ("+ds.label+") " \
		                                                                                                      "has been completed.\nMight be time to follow up with its owner, "+ds.owner.first_name+" "+ds.owner.last_name+ \
		             "("+ds.owner.username+")."
		html_content="<h4>Greetings!</h4> <p>A "+tasklabel+" reconciliation task for the dataset <b>"+ds.title+" ("+ds.label+")</b> " \
		                                                                                                                     "has been completed.</p><p>Might be time to follow up with its owner, "+ds.owner.first_name+" "+ds.owner.last_name+ \
		             " ("+ds.owner.username+").</p>"
		if task_name == 'wd':
			html_content += "<p>A nudge to mention that reconciling to the WHG index is helpful & worthwhile.</p>"
		elif task_name == 'idx':
			text_content = "Congratulations and thank you!\nYour *"+ds.title+"* dataset is now fully indexed \
      in World Historical Gazetteer. Where we already had one or more records for a place, yours is now linked to it/them.\n \
      For those we had no attestation for, yours is the new 'seed'. In any case, *all* your records are now accessible via \
      the index search, database search, and API.\nBest regards,\nThe WHG Team"
			html_content = "<h4>Congratulations and thank you!</h4><p>Your <b>"+ds.title+"</b> dataset is now fully indexed \ in World Historical Gazetteer. Where we already had one or more records for a place, yours is now linked to it/them.</p> \
      <p>For those we had no attestation for, yours is the new 'seed'. In any case, <i>all</i> your records are now accessible via \
      the index search, database search, and API.</p><p>Best regards,</p<p><i>The WHG Team</i></p>"
	except:
		print('status_emailer() failed on dsid', ds.id, 'how come?')
	subject, from_email = 'WHG dataset status update', settings.DEFAULT_FROM_EMAIL
	to_email = settings.EMAIL_STATUS_TO if task_name == 'wd' \
		else settings.EMAIL_STATUS_TO + [ds.owner.email]
	conn = mail.get_connection(
		host=settings.EMAIL_HOST,
		user=settings.EMAIL_HOST_USER,
		use_ssl=settings.EMAIL_USE_SSL,
		password=settings.EMAIL_HOST_PASSWORD,
		port=settings.EMAIL_PORT
	)
	msg = EmailMultiAlternatives(
		subject,
		text_content,
		from_email,
		to_email,
		connection=conn
	)
	msg.attach_alternative(html_content, "text/html")
	msg.send(fail_silently=False)

# TODO: faster?
class UpdateCountsView(View):
	""" Returns counts of unreviewed records, per pass and total; also deferred per task
	"""
	@staticmethod
	def get(request):
		#print('UpdateCountsView GET:',request.GET)
		"""
		args in request.GET:
				[integer] ds_id: dataset id
		"""
		ds = get_object_or_404(Dataset, id=request.GET.get('ds_id'))

		# deferred counts
		def defcountfunc(taskname, pids):
			if taskname[6:] in ['whg', 'idx']:
				return ds.places.filter(id__in=pids, review_whg = 2).count()
			elif taskname[6:].startswith('wd'):
				return ds.places.filter(id__in=pids, review_wd = 2).count()
			else:
				return ds.places.filter(id__in=pids, review_tgn = 2).count()

		def placecounter(th):
			pcounts={}
			#for th in taskhits.all():
			pcounts['p0'] = th.filter(query_pass='pass0').values('place_id').distinct().count()
			pcounts['p1'] = th.filter(query_pass='pass1').values('place_id').distinct().count()
			pcounts['p2'] = th.filter(query_pass='pass2').values('place_id').distinct().count()
			pcounts['p3'] = th.filter(query_pass='pass3').values('place_id').distinct().count()
			return pcounts

		updates={}
		# counts of distinct place ids w/unreviewed hits per task/pass
		# for t in ds.tasks.filter(status='SUCCESS'):
		#   taskhits = Hit.objects.filter(task_id=t.task_id, reviewed=False)
		for t in ds.tasks.filter(status='SUCCESS'):
			taskhits = Hit.objects.filter(task_id=t.task_id, reviewed=False)
			pcounts = placecounter(taskhits)
			# ids of all unreviewed places
			pids = list(set(taskhits.all().values_list("place_id",flat=True)))
			defcount = defcountfunc(t.task_name, pids)

			updates[t.task_id] = {
				"task":t.task_name,
				"total":len(pids),
				"pass0":pcounts['p0'],
				"pass1":pcounts['p1'],
				"pass2":pcounts['p2'],
				"pass3":pcounts['p3'],
				"deferred": defcount
			}

		#print(json.dumps(updates, indent=2))
		return JsonResponse(updates, safe=False)

# ***
# UPLOAD UTILS
# ***
def xl_tester():
	fn = '/Users/karlg/repos/_whgdata/data/_source/CentralEurasia/bregel_in progress.xlsx'
	from openpyxl import load_workbook
	wb = load_workbook(filename = fn)
	sheet_ranges = wb['range names']

def xl_upload(request):
	if "GET" == request.method:
		return render(request, 'datasets/xl.html', {})
	else:
		excel_file = request.FILES["excel_file"]

		# you may put validations here to check extension or file size

		wb = openpyxl.load_workbook(excel_file)

		# getting all sheets
		sheets = wb.sheetnames
		print(sheets)

		# getting a particular sheet by name out of many sheets
		ws = wb["Sheet1"]
		print(ws)

		excel_data = list()
		# iterating over the rows and
		# getting value from each cell in row
		for row in ws.iter_rows():
			row_data = list()
			for cell in row:
				#row_data.append(str(cell.value))
				row_data.append(cell.value)
			excel_data.append(row_data)

		return render(request, 'datasets/xl.html', {"excel_data":excel_data})


"""
	for hits in a wikidata task, create place_link and place_geom matches
	where geometry is within 5km
"""
def match_close_db(dsid, n=50000, test=True ):
  from datasets.models import Dataset, Hit
  from places.models import PlaceGeom, PlaceLink, Place
  from django.contrib.gis.geos import GEOSGeometry
  import json

  ds=Dataset.objects.get(id=dsid)
  task = ds.tasks.first()
  # hits = Hit.objects.filter(task_id = task.task_id, reviewed = False).order_by('id')
  hits = Hit.objects.filter(task_id = task.task_id,
                            reviewed = False,
                            ).order_by('place_id')
  count_add = 0
  geom_add = 0
  link_add = 0
  still_queued = 0
  no_geom = 0
  place_set = set([])

  for h in hits[:n]:
    place = Place.objects.get(id=h.place_id)
    pgs = [pg.geom for pg in place.geoms.all()]

		# skip if no geom in db place
    if len(pgs) == 0:
      # print('place id#'+str(place.id)+' ('+place.title+') has no geometry')
      no_geom +=1
      continue

    # skip if no geom in hit
    if 'geoms' in h.json and len(h.json) > 0:
	    qobj = {'pid': h.place_id, 'links': h.json['links'], 'geoms': h.json['geoms']}
    else:
      continue

    gwds = [GEOSGeometry(json.dumps(g)) for g in qobj['geoms']]
    qid = qobj['geoms'][0]['id']

    gwd = gwds[0]
    gpl = pgs[0] # first place geom

    try:
      dist = gwd.distance(gpl) * 100 # 3km more or less
    except:
      print('dist failed on', place, sys.exc_info())
      # print('gwd, gpl', gwd, gpl)
      continue
    # print(place.title, h.json['geoms'][0]['id'], dist)

		# add place_geom and place_link records if write==True
    if dist <= 5:
      # print('dist: ~' + str(dist)[:6] + 'km; '+str(h.authrecord_id), gwd.json)
      count_add +=1
      place_set.add(place.id)
      if not test:
        jsonb = {
          "type": gwd.geom_type,
          "citation": {"id": 'wd:' + qid, "label":'Wikidata' },
          "coordinates": qobj['geoms'][0]['coordinates']
        }

        # place, src_id, jsonb, task_id, geom
        pgobj = PlaceGeom.objects.create(
          place = place,
          geom = gwd,
          jsonb = jsonb,
          src_id = place.src_id,
          task_id = task.task_id
        )
        pgobj.save()
        geom_add +=1
        # place, src_id, jsonb, task_id
        # {"type": "closeMatch", "identifier": "tgn:7011198"}
        for i, l in enumerate(qobj['links']):
          # print('links to add:', i, l)
          link_add +=1
          plobj = PlaceLink.objects.create(
            place_id = place.id,
            src_id = place.src_id,
            task_id = task.task_id,
            jsonb={'type': 'closeMatch', 'identifier': l}
          )
          plobj.save()

        # flag hits and place
        place.review_wd = 1
        place.save()
        h.reviewed = True
        h.save()
    else:
      still_queued +=1
  # within 30km w/name match

  print("wrote match records" if not test else "only tested...")
  print("counts:", {"hits matched": count_add,
                    "places": len(list(place_set)),
                    "geoms": geom_add,
                    "no geom": no_geom,
                    "links": link_add,
                    "remaining": still_queued
                    })
